import openpyxl
from openpyxl.styles import Border, Side, Alignment, Font
from win32com.client.gencache import EnsureDispatch
from win32com.client import constants
import datetime
import pandas as pd
 
def df_to_html(df_with_dispo, df_without_dispo):
    '''
    Converts Pandas DataFrames extracted from extract_ncms_info to a formatted excel worksheet
    named "NCMS.xlsx" and then to an HTML page named "NCMS.htm"
 
        Args:
            df_with_dispo (DataFrame): DataFrame of NCRs with a disposition
            df_without_dispo (DataFrame): DataFrame of NCRs without a disposition
    '''
 
    filename = "C:/Users/B0640469/Downloads/NCMS.xlsx"
 
    NCRs_with_dispo = df_with_dispo[0]["NCR Number"].tolist()
 
    # add NCRs without a disposition to the DataFrame with dispositions
    for index, row in df_without_dispo[0].iterrows():
 
        if row["NCR Number"] not in NCRs_with_dispo:
            print(row["NCR Number"])
            modified_row = row.tolist()
            modified_row.insert(13, "PENDING DISPO")
            modified_row.insert(17,0)
            modified_row.insert(18,0)
            modified_row.insert(19,0)
            df_with_dispo[0].loc[len(df_with_dispo[0].index)] = modified_row
 
    # only keep most recent disposition for each unique NCR
    counts = df_with_dispo[0].groupby(['NCR Number', 'Discrepancy Number']).size().to_frame(name = 'Count').reset_index()
 
    df = pd.DataFrame(columns = df_with_dispo[0].columns.values)
 
    for index, row in counts.iterrows():
        NCR = row['NCR Number']
        discrepancy = row['Discrepancy Number']
        row_in_df = df_with_dispo[0].loc[(df_with_dispo[0]['NCR Number'] == NCR) & (df_with_dispo[0]['Discrepancy Number'] == discrepancy)]
        if row['Count'] > 1:
            df.loc[index] = row_in_df.iloc[-1]
        else:
            df.loc[index] = row_in_df.iloc[0]
 
    # sort DataFrame by NCR number
    df.sort_values(by='NCR Number')
 
    # load excel worksheet
    wb = openpyxl.load_workbook(filename)
    ws = wb.worksheets[0]
 
    # clear previous contents
    ws.delete_rows(5, ws.max_row)
 
    # add relevant information from DataFrame to each row of excel sheet
    for i in range(0, len(df)):
 
        ws.cell(i + 5, 1).value = i + 1
        ws.cell(i + 5, 4).value = 'D' + str(df.iloc[i][0])[6:] + '-' + str(df.iloc[i][1]) # NCR number
        #ws.cell(i + 5, 5).value = df.iloc[i][12] # discrepancy text
        ws.cell(i + 5, 6).value = df.iloc[i][14] # queue
        #ws.cell(i + 5, 7).value = df.iloc[i][13] # disposition text
 
        ws.row_dimensions[i].height = 20
 
    # add A/C number
    ws.cell(1, 1).value = "AC " + str(df.iloc[0]['Aircraft Number'])
 
    # add date and time of update
    ws.cell(2, 6).value = "DATE: " + str(datetime.date.today().strftime("%m-%d"))
    ws.cell(3, 6).value = "TIME: " + str(datetime.datetime.now().time().strftime("%H:%M:%S"))
 
    # add total number of NCRs
    ws.cell(3, 7).value = "TOTAL NCRs: " + str(ws.max_row - 4)
   
    # add borders and center text
    border = Border(left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin'))
 
    for row in ws['A5:M' + str(ws.max_row)]:
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.font = Font(size = "12")
 
    wb.save(filename)
   
    # save excel file as an HTML page
    xl = EnsureDispatch('Excel.Application')
    wb = xl.Workbooks.Open(filename)
    wb.SaveAs("C:/Users/B0640469/Downloads/NCMS.htm", constants.xlHtml)
    xl.Workbooks.Close()
    xl.Quit()